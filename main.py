import pyautogui
from openpyxl import load_workbook
import tkinter as tk
from tkinter import simpledialog, filedialog
from pynput.mouse import Listener, Button
import time


class ExcelDataProcessor:
    def __init__(self):
        self.file_path = None
        self.f_data = []
        self.i_data = []
        self.data_index = 0
        self.last_click_time = 0  # 将 last_click_time 作为类的属性


    def select_excel_file(self):
        root = tk.Tk()
        root.withdraw()
        self.file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx"), ("all files", "*.*"))
        )

    def get_column_ranges(self):
        root = tk.Tk()
        root.withdraw()
        start_f_col = simpledialog.askstring("Input", "学号起始单元格 (F10):")
        end_f_col = simpledialog.askstring("Input", "学号结束单元格 (FXX):")
        start_i_col = simpledialog.askstring("Input", "操行分起始单元格 (I10):")
        end_i_col = simpledialog.askstring("Input", "操行分结束单元格 (IXX):")
        return start_f_col, end_f_col, start_i_col, end_i_col

    def convert_excel_column(self, label):
        letter, number_str = label[:-2], label[-2:]
        number = int(number_str) - 1  # Convert to zero-based index
        return letter, number

    def read_excel_data(self, start_f_col, end_f_col, start_i_col, end_i_col):
        try:
            wb = load_workbook(self.file_path, data_only=True)
            ws = wb.active

            start_f_letter, start_f_number = self.convert_excel_column(start_f_col)
            end_f_letter, end_f_number = self.convert_excel_column(end_f_col)
            start_i_letter, start_i_number = self.convert_excel_column(start_i_col)
            end_i_letter, end_i_number = self.convert_excel_column(end_i_col)

            # Ensure the letters match and numbers are valid
            if (start_f_letter != end_f_letter or start_f_number >= ws.max_row or end_f_number >= ws.max_row or
                    start_i_letter != end_i_letter or start_i_number >= ws.max_row or end_i_number >= ws.max_row):
                raise ValueError("Invalid column range")

            f_data = [ws[f"{start_f_letter}{row}"].value for row in range(start_f_number + 1, end_f_number + 2)]
            i_data = [ws[f"{start_i_letter}{row}"].value for row in range(start_i_number + 1, end_i_number + 2)]
            self.f_data = f_data
            self.i_data = i_data
            print(f"First 10 rows of F data: {self.f_data[:10]}")
            print(f"First 10 rows of I data: {self.i_data[:10]}")
        except Exception as e:
            print(f"Error in read_excel_data: {e}")

    # @staticmethod
    def on_click(self, x, y, button, pressed):
        if not pressed and button == Button.left:
            current_time = time.time()
            if current_time - self.last_click_time < 0.3:  # 使用 self.last_click_time
                print(f'Double click detected at ({x}, {y})')
                if self.data_index < min(len(self.f_data), len(self.i_data)):
                    pyautogui.typewrite(str(self.f_data[self.data_index]))
                    print(f'F data: {self.f_data[self.data_index]} ')
                    pyautogui.press('tab')
                    time.sleep(0.1)
                    pyautogui.press('enter')
                    time.sleep(0.1)
                    pyautogui.typewrite(str(self.i_data[self.data_index]))
                    print(f'I data: {self.i_data[self.data_index]} ')
                    self.data_index += 1

            self.last_click_time = current_time  # 更新类的属性

    def run(self):
        self.select_excel_file()
        start_f_col, end_f_col, start_i_col, end_i_col = self.get_column_ranges()
        self.read_excel_data(start_f_col, end_f_col, start_i_col, end_i_col)
        with Listener(on_click=self.on_click) as listener:
            listener.join()


# 创建ExcelDataProcessor的实例
processor = ExcelDataProcessor()
processor.run()


