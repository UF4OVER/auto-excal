import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QWidget, QPushButton, QLabel, QFileDialog
import openpyxl
import os



class ExcelReaderApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.file_name = None


    def initUI(self):
        self.setWindowTitle('Excel File Reader')
        self.setGeometry(300, 300, 300, 200)

        central_widget = QWidget(self)
        self.setCentralWidget(central_widget)

        layout = QVBoxLayout(central_widget)

        self.file_label = QLabel('No file selected', self)
        layout.addWidget(self.file_label)

        self.data_label = QLabel('Data will be displayed here', self)
        layout.addWidget(self.data_label)

        select_button = QPushButton('Select Excel File', self)
        select_button.clicked.connect(self.selectFile)
        layout.addWidget(select_button)

    def selectFile(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_path, _ = QFileDialog.getOpenFileName(self, "Open Excel File", "", "Excel Files (*.xlsx)", options=options)
        if file_path:
            file_name = os.path.basename(file_path)  # 获取文件名
            self.file_name = file_name
            print("Selected file:", file_name)  # 打印文件名
            self.file_label.setText(file_name)  # 更新文件名标签
            data_f, data_h = self.readExcelData(file_path)  # 传入完整路径
            # 打印数据
            print("F Column Data:", data_f)
            print("H Column Data:", data_h)

    def readExcelData(self, file_path):
        if file_path is None:
            print("Error: No file path specified.")
            return [], []

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            data_f = []
            data_h = []

            for row in sheet.iter_rows(min_row=10, min_col=6, max_col=6, values_only=True):
                data_f.append(row[0])

            for row in sheet.iter_rows(min_row=10, min_col=9, max_col=9, values_only=True):
                data_h.append(row[0])

            return data_f, data_h

        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return [], []

#
# if __name__ == '__main__':
#     app = QApplication(sys.argv)
#     ex = ExcelReaderApp()
#     ex.show()
#     sys.exit(app.exec_())
#
